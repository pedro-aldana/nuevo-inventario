from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth.decorators import login_required

# autorizacion
from django.contrib.auth.decorators import user_passes_test

#----------


from django.http import HttpResponse
from django.contrib import messages
from .models import Producto,Mesa,Venta_mesa,Registro_ventas
from .forms import ProductForm,VentasForm,MesaForm
import openpyxl
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from django.db.models import Sum
# Create your views here.

# Reporte


def es_superusuario(user):
    return user.is_superuser

def generar_reporte(request):
    ventas = Registro_ventas.objects.all()
    
   # Crear un nuevo archivo de Excel
    wb = openpyxl.Workbook()
    wb.iso_dates = True
    # Obtener la hoja de trabajo activa
    sheet = wb.active

    # Agregar encabezados de columnas
    sheet['A1'] = 'Fecha'
    sheet['B1'] = 'Producto'
    sheet['C1'] = 'Cantidad'
    sheet['D1'] = '$ Total'

    # Agregar datos de ventas en las filas
    for index, venta in enumerate(ventas, start=2):
        sheet.cell(row=index, column=1).value = venta.fecha
        sheet.cell(row=index, column=2).value = venta.nombre
        sheet.cell(row=index, column=3).value = venta.cantidad_v
        sheet.cell(row=index, column=4).value = venta.monto

    # Formatear los números de la columna de Total con decimales después de la coma
    for row in sheet.iter_rows(min_row=2, max_row=len(ventas)+1, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = numbers.FORMAT_NUMBER_00
            
            
    # Formatear las celdas de la columna de Fecha como fecha
    for row in sheet.iter_rows(min_row=2, max_row=len(ventas)+1, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = 'yyyy-mm-dd'        
            

    # Configurar el nombre del archivo de descarga
    filename = 'reporte_ventas.xlsx'

    # Guardar el archivo de Excel en memoria
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response


# vISTAS DE PRODUCTOS

def Productos(request):
    
    productos = Producto.objects.filter(nombre__contains=request.GET.get('search', ''))
    
    context = {
        'productos': productos,
        
    }
    
    
    return render(request, 'pages/productos.html', context)



@login_required
def agregar(request):
    
    if request.method == 'POST':
        form = ProductForm(request.POST)
        
        if form.is_valid():
            form.save()
        
        return redirect('inventory:productos')
    
    else:    
    
        form = ProductForm()
        context = {
            'form':form
        }
        return render(request, 'pages/crud/añadir.html', context)

    
@user_passes_test(lambda u: u.is_superuser, login_url='/acceso_denegado/')
def edit(request,id):
    
    producto = Producto.objects.get(id=id)
    
    if request.method == 'GET':
        form = ProductForm(instance= producto)
        context ={
            'form':form,
            'id': id
        }
        return render(request, 'pages/crud/editar.html', context)
        
    if request.method == 'POST':
        form = ProductForm(request.POST, instance= producto)
        if form.is_valid():
            form.save()
        
        context = {
            'form':form,
            'id':id
        }
        messages.success(request, 'Producto actualizado correctamente')
        return render(request, 'pages/crud/editar.html',context) 
    


@user_passes_test(lambda u: u.is_superuser, login_url='/acceso_denegado/')    
def delete_2(request,id):
    producto = Producto.objects.get(id=id)
    producto.delete()
    return redirect('inventory:productos') 
    

# VISTAS DE VENTAS
@login_required
def ventas(request):
    
    ventas = Registro_ventas.objects.filter(nombre__contains=request.GET.get('search', ''))
    
    context = {
        'ventas':ventas
    }
    
    return render(request, 'pages/ventas.html', context)

@login_required
def vender(request):
    try:
    
        if request.method == 'POST':
            form = VentasForm(request.POST)
            
            if form.is_valid():
                form.save()
            
            return redirect('inventory:mesas')
        
        else:
            
            form = VentasForm()
            context = {
                'form':form
            }
            return render(request, 'pages/crud/vender.html', context)        
    except ValueError as e:
        error_message = str(e)
        return  render(request, 'messages/error.html', {
            'error_message': error_message
        })   

@login_required
def edit_ventas(request,id):
    
    venta = Venta_mesa.objects.get(id=id)
    
    if request.method == 'GET':
        form = VentasForm(instance= venta)
        context ={
            'form':form,
            'id': id
        }
        return render(request, 'pages/crud/editar_venta.html', context)
        
    if request.method == 'POST':
        form = VentasForm(request.POST, instance= venta)
        if form.is_valid():
            form.save()
        
        context = {
            'form':form,
            'id':id
        }
        messages.success(request, 'Venta Editada correctamente')
        return render(request, 'pages/crud/editar_venta.html',context) 
    
    
# VISTAS DE MESAS
@login_required
def mesas(request):
    
    mesa = Mesa.objects.all()
    
    context = {
        'mesas':mesa
    }
    return render(request, 'pages/mesas.html',context)

@login_required
def agregar_mesas(request):
    
    if request.method == 'POST':
        form = MesaForm(request.POST)
        
        if form.is_valid():
            form.save()
        
        return redirect('inventory:mesas')
    
    else:    
    
        form = MesaForm()
        context = {
            'form':form
        }
        return render(request, 'pages/crud/ag_mesa.html',context)
    
    

@login_required
def detail(request, mesa_id):
    
    mesa = get_object_or_404(Mesa, id=mesa_id)
    ventas = Venta_mesa.objects.filter(mesaId=mesa)
    total_pagar_suma = ventas.aggregate(total_suma=Sum('monto_mesa'))['total_suma']
    
    

    context = {
        'mesa': mesa,
        'ventas':ventas,
        'total_pagar_suma':total_pagar_suma
        
    }
    
    return render(request, 'pages/detail.html',context)    

@login_required
def delete(request,id):
    venta_mesa = Venta_mesa.objects.get(id=id)
    venta_mesa.delete()
    return redirect('inventory:mesas')

def delete_mesa(request,id):
    mesa = Mesa.objects.get(id=id)
    mesa.delete()
    return redirect('inventory:mesas')


# acceso denegado

def acceso_denegado(request):
    mensaje = "Lo siento, no tienes acceso a esta página."
    return render(request, 'messages/acceso_denegado.html', {'mensaje': mensaje})